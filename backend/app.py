from __future__ import annotations

from pathlib import Path
from typing import Any
from datetime import datetime, timezone
import io

import joblib
import pandas as pd
from flask import Flask, jsonify, request, send_from_directory, make_response
from sklearn.preprocessing import LabelEncoder
import logging
from fpdf import FPDF, XPos, YPos
from openpyxl import Workbook

# ================== Setup ==================
app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent
FRONTEND_DIR = PROJECT_ROOT / "frontend"

FEATURE_COLUMNS = [
    "strength_rating",
    "weight_capacity",
    "biodegradability_score",
    "recyclability_percentage",
    "material_type_encoded",
]

REQUIRED_INPUT_FIELDS = [
    "strength_rating",
    "weight_capacity",
    "biodegradability_score",
    "recyclability_percentage",
    "material_type",
]

LEGACY_ENGINEERED_FEATURES = [
    "cost_efficiency_index",
    "co2_impact_index",
    "material_suitability_score",
]

# ================== CORS ==================
@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response

# ================== Helpers ==================
def _find_existing_path(candidates: list[Path], label: str) -> Path:
    for candidate in candidates:
        if candidate.exists() and candidate.stat().st_size > 0:
            return candidate
    checked = "\n".join(str(path) for path in candidates)
    raise FileNotFoundError(f"Could not find a valid {label}. Checked:\n{checked}")

def _add_engineered_features(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()

    if "material_suitability_score" not in result.columns:
        result["material_suitability_score"] = (
            0.4 * result["strength_rating"]
            + 0.3 * result["biodegradability_score"]
            + 0.2 * result["recyclability_percentage"]
            - 0.1 * result.get("co2_emission_score", 0.0)
        )

    if "cost_efficiency_index" not in result.columns:
        proxy_cost = (
            1.0
            + 0.08 * result["weight_capacity"]
            + 0.20 * (11.0 - result["biodegradability_score"])
        )
        result["cost_efficiency_index"] = result["strength_rating"] / proxy_cost.replace(0, 1.0)

    if "co2_impact_index" not in result.columns:
        proxy_co2 = (
            0.60 * result["strength_rating"]
            + 0.08 * result["weight_capacity"]
            - 0.05 * result["recyclability_percentage"]
            - 0.25 * result["biodegradability_score"]
            + 6.0
        )
        result["co2_impact_index"] = (
            proxy_co2.clip(lower=0.1) * 10 / result["recyclability_percentage"].clip(lower=1.0)
        )

    return result

def _load_resources() -> dict[str, Any]:
    model_candidates = [BASE_DIR / "models", PROJECT_ROOT / "ml" / "models"]

    cost_model_path = _find_existing_path(
        [path / "cost_model.pkl" for path in model_candidates],
        "cost model",
    )
    co2_model_path = _find_existing_path(
        [path / "co2_model.pkl" for path in model_candidates],
        "CO2 model",
    )
    scaler_path = _find_existing_path(
        [path / "scaler.pkl" for path in model_candidates],
        "scaler",
    )

    dataset_path = _find_existing_path(
        [
            BASE_DIR / "dataset" / "engineered_materials.csv",
            PROJECT_ROOT / "ml" / "dataset" / "engineered_materials.csv",
        ],
        "engineered dataset",
    )

    cost_model = joblib.load(cost_model_path)
    co2_model = joblib.load(co2_model_path)
    scaler = joblib.load(scaler_path)

    dataset = pd.read_csv(dataset_path)
    if dataset.empty:
        raise ValueError(f"Dataset is empty: {dataset_path}")

    if "material_type" not in dataset.columns:
        raise ValueError("Dataset is missing required column: material_type")

    encoder = LabelEncoder()
    dataset["material_type_encoded"] = encoder.fit_transform(dataset["material_type"].astype(str))
    dataset = _add_engineered_features(dataset)

    feature_columns = list(getattr(scaler, "feature_names_in_", FEATURE_COLUMNS))
    for column in feature_columns:
        if column not in dataset.columns:
            raise ValueError(f"Dataset is missing feature column: {column}")

    logging.info(f"Loaded resources: {dataset_path}, models, scaler")
    return {
        "cost_model": cost_model,
        "co2_model": co2_model,
        "scaler": scaler,
        "dataset": dataset,
        "encoder": encoder,
        "feature_columns": feature_columns,
        "paths": {
            "cost_model": str(cost_model_path),
            "co2_model": str(co2_model_path),
            "scaler": str(scaler_path),
            "dataset": str(dataset_path),
        },
    }

def _validate_and_prepare_materials(materials: list[dict[str, Any]], encoder: LabelEncoder) -> pd.DataFrame:
    if not materials:
        raise ValueError("materials must be a non-empty list")

    frame = pd.DataFrame(materials)
    missing_columns = [field for field in REQUIRED_INPUT_FIELDS if field not in frame.columns]
    if missing_columns:
        raise ValueError(f"Missing required fields: {', '.join(missing_columns)}")

    for numeric_col in REQUIRED_INPUT_FIELDS:
        if numeric_col != "material_type":
            frame[numeric_col] = pd.to_numeric(frame[numeric_col], errors="coerce")

    if frame.drop(columns=["material_type"]).isnull().any().any():
        raise ValueError("One or more numeric fields have invalid values")

    unknown_types = sorted(set(frame["material_type"].astype(str)) - set(encoder.classes_))
    if unknown_types:
        raise ValueError(
            f"Unknown material_type values: {', '.join(unknown_types)}. Allowed: {', '.join(map(str, encoder.classes_))}"
        )

    frame["material_type_encoded"] = encoder.transform(frame["material_type"].astype(str))
    return _add_engineered_features(frame)

def _score_materials(frame: pd.DataFrame, resources: dict[str, Any], top_n: int | None = None, filter_type: str | None = None) -> pd.DataFrame:
    model_frame = frame.copy()
    
    if filter_type:
        model_frame = model_frame[model_frame["material_type"] == filter_type]
    if model_frame.empty:
        return model_frame

    X = model_frame[resources["feature_columns"]]
    X_scaled = resources["scaler"].transform(X)

    model_frame["predicted_cost"] = resources["cost_model"].predict(X_scaled)
    model_frame["predicted_co2"] = resources["co2_model"].predict(X_scaled)

    model_frame["eco_score"] = (
        0.4 * model_frame["biodegradability_score"]
        + 0.3 * model_frame["recyclability_percentage"]
        - 0.2 * model_frame["predicted_co2"]
        - 0.1 * model_frame["predicted_cost"]
    )
    model_frame = model_frame.sort_values(by="eco_score", ascending=False).reset_index(drop=True)

    if top_n:
        model_frame = model_frame.head(top_n)
    return model_frame

def _safe_percent_delta(reference: float, candidate: float, higher_is_better: bool = False) -> float:
    base = abs(float(reference))
    if base <= 1e-9:
        return 0.0
    if higher_is_better:
        return (float(candidate) - float(reference)) / base * 100.0
    return (float(reference) - float(candidate)) / base * 100.0

def _join_phrases(phrases: list[str]) -> str:
    cleaned = [phrase.strip() for phrase in phrases if phrase and phrase.strip()]
    if not cleaned:
        return ""
    if len(cleaned) == 1:
        return cleaned[0]
    if len(cleaned) == 2:
        return f"{cleaned[0]} and {cleaned[1]}"
    return f"{', '.join(cleaned[:-1])}, and {cleaned[-1]}"

def _safe_top_n(value: Any, default: int = 5, min_value: int = 1) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        parsed = default
    return max(parsed, min_value)

def _format_material(row: pd.Series) -> dict[str, Any]:
    durability_score = row.get("material_suitability_score", row.get("strength_rating", 0.0))
    return {
        "material_name": row.get("material_name", "Custom Material"),
        "material_type": row["material_type"],
        "strength_rating": round(float(row.get("strength_rating", 0.0)), 4),
        "weight_capacity": round(float(row.get("weight_capacity", 0.0)), 4),
        "biodegradability_score": round(float(row.get("biodegradability_score", 0.0)), 4),
        "recyclability_percentage": round(float(row.get("recyclability_percentage", 0.0)), 4),
        "durability_score": round(float(durability_score), 4),
        "predicted_cost": round(float(row["predicted_cost"]), 4),
        "predicted_co2": round(float(row["predicted_co2"]), 4),
        "eco_score": round(float(row["eco_score"]), 4),
        "final_ranking_score": round(float(row["eco_score"]), 4),
    }

def _build_material_badges(scored_frame: pd.DataFrame, index: int) -> list[str]:
    if scored_frame.empty:
        return []

    row = scored_frame.iloc[index]
    top_slice = scored_frame.head(min(3, len(scored_frame)))
    badges: list[str] = []

    if index == 0:
        badges.append("Top Ranked")
    if float(row["predicted_co2"]) <= float(top_slice["predicted_co2"].min()):
        badges.append("Lowest CO2")
    if float(row["recyclability_percentage"]) >= float(top_slice["recyclability_percentage"].max()):
        badges.append("Most Recyclable")

    row_efficiency = float(row["eco_score"]) / max(float(row["predicted_cost"]), 1e-6)
    top_efficiency = (top_slice["eco_score"] / top_slice["predicted_cost"].clip(lower=1e-6)).max()
    if row_efficiency >= float(top_efficiency):
        badges.append("Best Cost Balance")

    return badges[:3]

def _build_material_explanation(scored_frame: pd.DataFrame, index: int) -> dict[str, Any]:
    # Build explanation text from real ranking signals so frontend does not rely on hardcoded narratives.
    row = scored_frame.iloc[index]
    next_row = scored_frame.iloc[index + 1] if index + 1 < len(scored_frame) else None
    median_values = scored_frame[["predicted_co2", "predicted_cost", "strength_rating", "biodegradability_score", "recyclability_percentage"]].median()

    reasons: list[str] = []
    if float(row["predicted_co2"]) <= float(median_values["predicted_co2"]):
        reasons.append("lower predicted CO2 emissions")
    if float(row["predicted_cost"]) <= float(median_values["predicted_cost"]):
        reasons.append("competitive predicted cost")
    if float(row["strength_rating"]) >= float(median_values["strength_rating"]):
        reasons.append("strong durability for packaging use")
    if float(row["biodegradability_score"]) >= float(median_values["biodegradability_score"]):
        reasons.append("high biodegradability")
    if float(row["recyclability_percentage"]) >= float(median_values["recyclability_percentage"]):
        reasons.append("high recyclability")

    if len(reasons) < 2:
        reasons.extend(
            [
                "a balanced sustainability-cost profile",
                "consistent durability for transport and handling",
            ]
        )

    relative_comparison: list[str] = []
    if next_row is not None:
        co2_better = _safe_percent_delta(float(next_row["predicted_co2"]), float(row["predicted_co2"]))
        cost_better = _safe_percent_delta(float(next_row["predicted_cost"]), float(row["predicted_cost"]))
        eco_advantage = float(row["eco_score"]) - float(next_row["eco_score"])
        if abs(co2_better) >= 0.1:
            if co2_better >= 0:
                relative_comparison.append(f"{co2_better:.1f}% lower CO2 than rank {index + 2}.")
            else:
                relative_comparison.append(f"{abs(co2_better):.1f}% higher CO2 than rank {index + 2}.")
        if abs(cost_better) >= 0.1:
            if cost_better >= 0:
                relative_comparison.append(f"{cost_better:.1f}% lower predicted cost than rank {index + 2}.")
            else:
                relative_comparison.append(f"{abs(cost_better):.1f}% higher predicted cost than rank {index + 2}.")
        if abs(eco_advantage) >= 0.01:
            if eco_advantage >= 0:
                relative_comparison.append(f"Eco score lead of {eco_advantage:.2f} over rank {index + 2}.")
            else:
                relative_comparison.append(f"Eco score trails rank {index + 2} by {abs(eco_advantage):.2f}.")

    plastic_group = scored_frame[
        scored_frame["material_type"].astype(str).str.contains("plastic", case=False, na=False)
    ]
    if not plastic_group.empty:
        row_efficiency = float(row["eco_score"]) / max(float(row["predicted_cost"]), 1e-6)
        plastic_efficiency = (
            plastic_group["eco_score"] / plastic_group["predicted_cost"].clip(lower=1e-6)
        ).mean()
        efficiency_gain = _safe_percent_delta(
            float(plastic_efficiency), float(row_efficiency), higher_is_better=True
        )
        if abs(efficiency_gain) >= 0.1:
            if efficiency_gain >= 0:
                relative_comparison.append(
                    f"{efficiency_gain:.1f}% more cost-efficient than plastic-based alternatives."
                )
            else:
                relative_comparison.append(
                    f"{abs(efficiency_gain):.1f}% less cost-efficient than plastic-based alternatives."
                )

    primary_reason = _join_phrases(reasons[:3])
    why_selected = (
        f"This material is recommended because it delivers {primary_reason} while maintaining "
        f"an overall ranking score of {float(row['eco_score']):.2f}."
    )

    return {
        "why_selected": why_selected,
        "relative_comparison": relative_comparison,
        "key_factors": reasons[:4],
    }

def _serialize_ranked_materials(scored_frame: pd.DataFrame) -> list[dict[str, Any]]:
    if scored_frame.empty:
        return []

    ranked_output: list[dict[str, Any]] = []
    for index, (_, row) in enumerate(scored_frame.iterrows()):
        material = _format_material(row)
        material["rank"] = index + 1
        material["badges"] = _build_material_badges(scored_frame, index)
        material["explanation"] = _build_material_explanation(scored_frame, index)
        ranked_output.append(material)
    return ranked_output

def _build_rank_winner_reason(top_materials: list[dict[str, Any]]) -> str:
    if not top_materials:
        return "No ranked materials were available for comparison."
    if len(top_materials) == 1:
        return "Only one candidate was available, so it is selected by default."

    first = top_materials[0]
    second = top_materials[1]
    co2_better = _safe_percent_delta(second["predicted_co2"], first["predicted_co2"])
    cost_better = _safe_percent_delta(second["predicted_cost"], first["predicted_cost"])
    eco_advantage = first["eco_score"] - second["eco_score"]
    co2_text = (
        f"{abs(co2_better):.1f}% lower CO2" if co2_better >= 0 else f"{abs(co2_better):.1f}% higher CO2"
    )
    cost_text = (
        f"{abs(cost_better):.1f}% lower predicted cost"
        if cost_better >= 0
        else f"{abs(cost_better):.1f}% higher predicted cost"
    )
    eco_text = (
        f"+{eco_advantage:.2f} eco-score advantage"
        if eco_advantage >= 0
        else f"{abs(eco_advantage):.2f} eco-score disadvantage"
    )

    return (
        f"{first['material_name']} compared to rank 2: {co2_text}, {cost_text}, and a {eco_text}."
    )

def _summarize_by_material_type(scored_frame: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        scored_frame.groupby("material_type")
        .agg(
            count=("material_type", "size"),
            avg_cost=("predicted_cost", "mean"),
            avg_co2=("predicted_co2", "mean"),
            avg_eco=("eco_score", "mean"),
        )
        .sort_values("count", ascending=False)
        .reset_index()
    )
    return grouped

def _build_actionable_insights(scored: pd.DataFrame, baseline_cost: float, baseline_co2: float) -> list[dict[str, Any]]:
    if scored.empty:
        return []

    # Keep dashboard insights business-friendly while grounding every card in model predictions.
    best = scored.iloc[0]
    best_cost = float(best["predicted_cost"])
    best_co2 = float(best["predicted_co2"])

    co2_reduction_vs_baseline = _safe_percent_delta(baseline_co2, best_co2)
    cost_savings_vs_baseline = baseline_cost - best_cost

    low_cost_eco = scored[
        (scored["predicted_cost"] <= baseline_cost)
        & (scored["eco_score"] >= scored["eco_score"].quantile(0.70))
    ]
    if low_cost_eco.empty:
        low_cost_eco = scored.nsmallest(1, "predicted_cost")
    else:
        low_cost_eco = low_cost_eco.sort_values(["predicted_cost", "predicted_co2"]).head(1)
    low_cost_eco_row = low_cost_eco.iloc[0]

    return [
        {
            "title": "Recommended switch",
            "value": best.get("material_name", "Top Material"),
            "insight": (
                f"Switching to {best.get('material_name', 'this material')} could reduce CO2 by "
                f"{co2_reduction_vs_baseline:.1f}% compared to current dataset baseline."
            ),
            "action": "Pilot this material in the next procurement cycle.",
            "badge": "Best Choice",
        },
        {
            "title": "Estimated packaging cost savings",
            "value": round(cost_savings_vs_baseline, 4),
            "insight": (
                f"Estimated savings of {cost_savings_vs_baseline:.2f} per unit versus average baseline cost."
            ),
            "action": "Prioritize this option for high-volume SKUs first.",
            "badge": "Cost Impact",
        },
        {
            "title": "Most sustainable material",
            "value": best.get("material_name", "Top Material"),
            "insight": (
                f"Highest eco score in the dataset: {float(best['eco_score']):.2f} with predicted CO2 of "
                f"{best_co2:.2f}."
            ),
            "action": "Use as default where compliance requires strongest sustainability metrics.",
            "badge": "Sustainability Leader",
        },
        {
            "title": "Best low-cost eco alternative",
            "value": low_cost_eco_row.get("material_name", "Alternative"),
            "insight": (
                f"{low_cost_eco_row.get('material_name', 'This option')} balances low cost "
                f"({float(low_cost_eco_row['predicted_cost']):.2f}) with eco score "
                f"{float(low_cost_eco_row['eco_score']):.2f}."
            ),
            "action": "Use this option when budget is the primary constraint.",
            "badge": "Budget Friendly",
        },
    ]

def _build_dashboard_summary(resources: dict[str, Any], top_n: int = 5, filter_type: str | None = None) -> dict[str, Any]:
    scored = _score_materials(resources["dataset"], resources, top_n=None, filter_type=filter_type)
    if scored.empty:
        raise ValueError("No materials available for dashboard summary")

    safe_top_n = max(int(top_n or 5), 1)
    safe_top_n = min(safe_top_n, len(scored))

    baseline_cost = float(scored["predicted_cost"].mean())
    baseline_co2 = float(scored["predicted_co2"].mean())

    top_frame = scored.head(safe_top_n)
    top_cost = float(top_frame["predicted_cost"].mean())
    top_co2 = float(top_frame["predicted_co2"].mean())
    top_eco = float(top_frame["eco_score"].mean())

    cost_savings_pct = ((baseline_cost - top_cost) / baseline_cost * 100.0) if baseline_cost > 0 else 0.0
    co2_reduction_pct = ((baseline_co2 - top_co2) / baseline_co2 * 100.0) if baseline_co2 > 0 else 0.0

    usage = _summarize_by_material_type(scored).head(8)
    ranked_top = _serialize_ranked_materials(top_frame)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "filter_type": filter_type,
        "baseline": {"avg_cost": round(baseline_cost, 4), "avg_co2": round(baseline_co2, 4)},
        "top_summary": {
            "top_n": safe_top_n,
            "avg_cost": round(top_cost, 4),
            "avg_co2": round(top_co2, 4),
            "avg_eco_score": round(top_eco, 4),
        },
        "savings": {
            "cost_savings_pct": round(cost_savings_pct, 2),
            "co2_reduction_pct": round(co2_reduction_pct, 2),
        },
        "usage_trends": {
            "labels": usage["material_type"].tolist(),
            "counts": usage["count"].round(0).astype(int).tolist(),
            "avg_cost": usage["avg_cost"].round(4).tolist(),
            "avg_co2": usage["avg_co2"].round(4).tolist(),
            "avg_eco": usage["avg_eco"].round(4).tolist(),
        },
        "best_material": ranked_top[0],
        "top_materials": ranked_top,
        "actionable_insights": _build_actionable_insights(scored, baseline_cost, baseline_co2),
    }

def _render_pdf_report(summary: dict[str, Any]) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "EcoPackAI Sustainability Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, f"Generated (UTC): {summary['generated_at']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    if summary.get("filter_type"):
        pdf.cell(
            0,
            8,
            f"Filtered Material Type: {summary['filter_type']}",
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
        )
    pdf.ln(4)

    baseline = summary["baseline"]
    top = summary["top_summary"]
    savings = summary["savings"]

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Key Metrics", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 7, f"Baseline Avg Cost: {baseline['avg_cost']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"Baseline Avg CO2: {baseline['avg_co2']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"Top {top['top_n']} Avg Cost: {top['avg_cost']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"Top {top['top_n']} Avg CO2: {top['avg_co2']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"Cost Savings %: {savings['cost_savings_pct']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 7, f"CO2 Reduction %: {savings['co2_reduction_pct']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Top Recommended Materials", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=10)

    col_widths = [50, 30, 30, 30, 30]
    headers = ["Material", "Type", "Eco", "Cost", "CO2"]
    for header, width in zip(headers, col_widths):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()

    for material in summary["top_materials"]:
        row = [
            material["material_name"],
            material["material_type"],
            f"{material['eco_score']:.3f}",
            f"{material['predicted_cost']:.2f}",
            f"{material['predicted_co2']:.2f}",
        ]
        for value, width in zip(row, col_widths):
            pdf.cell(width, 7, str(value)[:24], border=1)
        pdf.ln()

    # fpdf2 may return `str`, `bytes`, or `bytearray` depending on version.
    pdf_output = pdf.output()
    if isinstance(pdf_output, str):
        return pdf_output.encode("latin-1")
    return bytes(pdf_output)

def _render_excel_report(summary: dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    summary_ws.append(["Metric", "Value"])

    baseline = summary["baseline"]
    top = summary["top_summary"]
    savings = summary["savings"]

    summary_ws.append(["Generated (UTC)", summary["generated_at"]])
    summary_ws.append(["Baseline Avg Cost", baseline["avg_cost"]])
    summary_ws.append(["Baseline Avg CO2", baseline["avg_co2"]])
    summary_ws.append([f"Top {top['top_n']} Avg Cost", top["avg_cost"]])
    summary_ws.append([f"Top {top['top_n']} Avg CO2", top["avg_co2"]])
    summary_ws.append(["Cost Savings %", savings["cost_savings_pct"]])
    summary_ws.append(["CO2 Reduction %", savings["co2_reduction_pct"]])

    if summary.get("filter_type"):
        summary_ws.append(["Filtered Material Type", summary["filter_type"]])

    top_ws = workbook.create_sheet("Top Materials")
    top_ws.append(["Material Name", "Type", "Eco Score", "Predicted Cost", "Predicted CO2"])
    for material in summary["top_materials"]:
        top_ws.append(
            [
                material["material_name"],
                material["material_type"],
                material["eco_score"],
                material["predicted_cost"],
                material["predicted_co2"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()

def _build_file_download_response(content: bytes, mime_type: str, filename: str):
    response = make_response(content)
    response.headers["Content-Type"] = mime_type
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.headers["Content-Length"] = str(len(content))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    return response

# ================== Load Resources ==================
try:
    RESOURCES = _load_resources()
    STARTUP_ERROR = None
except Exception as exc:
    RESOURCES = None
    STARTUP_ERROR = str(exc)
    logging.error(f"Startup failed: {STARTUP_ERROR}")

@app.get("/")
def home() -> Any:
    return send_from_directory(FRONTEND_DIR, "index.html")

@app.get("/style.css")
def frontend_css() -> Any:
    return send_from_directory(FRONTEND_DIR, "style.css")

@app.get("/script.js")
def frontend_js() -> Any:
    return send_from_directory(FRONTEND_DIR, "script.js")

@app.get("/frontend/<path:filename>")
def frontend_assets(filename: str) -> Any:
    return send_from_directory(FRONTEND_DIR, filename)

@app.get("/health")
def health() -> Any:
    if STARTUP_ERROR:
        return jsonify({"status": "error", "message": STARTUP_ERROR}), 500
    dataset = RESOURCES["dataset"]
    return jsonify(
        {
            "status": "ok",
            "dataset_rows": int(len(dataset)),
            "dataset_columns": list(dataset.columns),
            "feature_columns": RESOURCES["feature_columns"],
            "resources": RESOURCES["paths"],
        }
    )

@app.get("/metadata/material-types")
def material_types() -> Any:
    if STARTUP_ERROR:
        return jsonify({"error": STARTUP_ERROR}), 500
    classes = sorted(str(value) for value in RESOURCES["encoder"].classes_)
    return jsonify({"material_types": classes, "count": len(classes)})

@app.get("/recommend")
def recommend_from_dataset() -> Any:
    if STARTUP_ERROR:
        return jsonify({"error": STARTUP_ERROR}), 500

    requested_top_n = request.args.get("top_n", default=5, type=int)
    filter_type = request.args.get("material_type", default=None, type=str)
    top_n = _safe_top_n(requested_top_n, default=5, min_value=3)

    ranked = _score_materials(RESOURCES["dataset"], RESOURCES, top_n=top_n, filter_type=filter_type)
    if ranked.empty:
        return jsonify({"error": "No materials available for the selected filter."}), 400
    ranked_materials = _serialize_ranked_materials(ranked)
    best = ranked_materials[0]
    top_3 = ranked_materials[: min(3, len(ranked_materials))]

    return jsonify({
        "source": "dataset",
        "best_material": best,
        "top_ranked": ranked_materials,
        "ranked_materials": ranked_materials,
        "top_3_comparison": {
            "materials": top_3,
            "why_rank_1_wins": _build_rank_winner_reason(top_3),
        },
    })

@app.post("/recommend")
def recommend_from_payload() -> Any:
    if STARTUP_ERROR:
        return jsonify({"error": STARTUP_ERROR}), 500

    payload = request.get_json(silent=True) or {}
    materials = payload.get("materials")
    requested_top_n = payload.get("top_n", None)
    filter_type = payload.get("material_type", None)
    top_n = _safe_top_n(requested_top_n, default=5, min_value=3)

    if not isinstance(materials, list):
        return jsonify({"error": "Request body must include 'materials' as a list"}), 400

    try:
        prepared = _validate_and_prepare_materials(materials, RESOURCES["encoder"])
        ranked = _score_materials(prepared, RESOURCES, top_n=top_n, filter_type=filter_type)
        if ranked.empty:
            raise ValueError("No materials available for the selected filter.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logging.exception("Failed to score materials")
        return jsonify({"error": f"Failed to score materials: {exc}"}), 500

    ranked_materials = _serialize_ranked_materials(ranked)
    best = ranked_materials[0]
    top_3 = ranked_materials[: min(3, len(ranked_materials))]

    return jsonify({
        "source": "request_payload",
        "best_material": best,
        "ranked_materials": ranked_materials,
        "top_ranked": ranked_materials,
        "top_3_comparison": {
            "materials": top_3,
            "why_rank_1_wins": _build_rank_winner_reason(top_3),
        },
    })

@app.get("/analytics/summary")
def analytics_summary() -> Any:
    if STARTUP_ERROR:
        return jsonify({"error": STARTUP_ERROR}), 500

    top_n = request.args.get("top_n", default=5, type=int)
    filter_type = request.args.get("material_type", default=None, type=str)

    try:
        summary = _build_dashboard_summary(RESOURCES, top_n=top_n, filter_type=filter_type)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logging.exception("Failed to build dashboard summary")
        return jsonify({"error": f"Failed to build dashboard summary: {exc}"}), 500

    return jsonify(summary)

@app.get("/reports/sustainability")
def sustainability_report() -> Any:
    if STARTUP_ERROR:
        return jsonify({"error": STARTUP_ERROR}), 500

    report_format = request.args.get("format", default="pdf", type=str).lower()
    top_n = request.args.get("top_n", default=5, type=int)
    filter_type = request.args.get("material_type", default=None, type=str)

    try:
        summary = _build_dashboard_summary(RESOURCES, top_n=top_n, filter_type=filter_type)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        logging.exception("Failed to build report summary")
        return jsonify({"error": f"Failed to build report summary: {exc}"}), 500

    if report_format in {"pdf"}:
        content = _render_pdf_report(summary)
        return _build_file_download_response(
            content=content,
            mime_type="application/pdf",
            filename="EcoPackAI_Sustainability_Report.pdf",
        )
    if report_format in {"excel", "xlsx"}:
        content = _render_excel_report(summary)
        return _build_file_download_response(
            content=content,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="EcoPackAI_Sustainability_Report.xlsx",
        )

    return jsonify({"error": "Unsupported format. Use pdf or excel."}), 400

# ================== Run Server ==================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
